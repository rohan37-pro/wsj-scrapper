from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

import requests
import json
import sys
import os
import openpyxl
from time import sleep

working_directory = os.getcwd()

def credentials():
	with open('credentials.txt','r') as file:
		data = file.read()
		data = json.loads(data)
		username = data['user_name'].strip()
		password  = data["password"].strip()
	return username , password


def login():
	options = webdriver.ChromeOptions()
	options.add_argument('--user-data-dir=./user')
	driver = webdriver.Chrome("./chromedriver",options=options)
	driver.get("https://accounts.wsj.com/login")
	sleep(3)

	username, password  = credentials()
	print(username, password)

	element = driver.find_element('xpath', "(//div[@class='fancy-text-input'])[1]//input")
	element.send_keys(username)
	action = ActionChains(driver)
	action.send_keys(Keys.ENTER).perform()
	sleep(3)
	action.send_keys(password).send_keys(Keys.ENTER).perform()
	sleep(5)

	element = driver.find_element('xpath', "//button[@class='solid-button reg-rtl-btn']")
	element.click()

	sleep(2)
	driver.quit()



def main():
	try:
		if sys.argv[1] == 'login':
			login()
	except:
		pass

	options = webdriver.ChromeOptions()
	options.add_argument('--user-data-dir=./user')

	driver = webdriver.Chrome("./chromedriver",options=options)

	driver.get("https://www.wsj.com/news/types/crypto?mod=breadcrumb")
	sleep(2)

	article_list = []
	for i in range(1,4):
	
		stories = driver.find_element('xpath',f"(//div[@id='latest-stories']//article[{i}]//a)[2]")
		author = driver.find_element('xpath',f"(//div[@id='latest-stories']//article[{i}]//p)[2]")
		date = driver.find_element('xpath',f"(//div[@id='latest-stories']//article[{i}]//p)[3]")
		image = driver.find_element('xpath',f"(//div[@id='latest-stories']//article[{i}]//img)")
		head_line = driver.find_element('xpath',f"(//div[@id='latest-stories']//article[{i}]//span)[1]")

		author = author.text
		date = date.text
		head_line = head_line.text
		image = image.get_attribute('src')
		article = stories.get_attribute("href")
		article_list.append(article)

		print(f"author : {author}")
		print(f"date : {date}")
		print(f"image : {image}")
		print(f"article : {article}")
		print(f"head_line : {head_line}")
		print()

		# image = requests.get(image)
		all_data_of_articles.append([head_line, author, date, image])



	i = 1
	for article in article_list:
		add_para = ''
		driver.get(article)
		sleep(2)

		head_line = all_data_of_articles[i][0]
		author = all_data_of_articles[i][1]
		date = all_data_of_articles[i][2]
		image = driver.find_element('xpath', "(//div[@class='paywall css-1u1nl00-PaywalledContentContainer e1db8bjv0']//img)[1]")
		image_link = image.get_attribute('srcset')
		image_link = image_link.split(" ")
		image_link = image_link[-2]
		print(f"image_link ------->    {image_link}")
		image = requests.get(image_link)
		image = image.content
		# video = driver.find_element('xpath', "(//div[@class='video-wrapper player-8U']//video)[1]")
		# video.a
		paragraphs = driver.find_elements('xpath', "//section//p[@data-type='paragraph']")

		for p in paragraphs:
			add_para += p.text + '\n'

		print(add_para)
		all_data_of_articles[i].append(add_para)

		add_para = f'''{head_line}\n\n\n{author}\n{date}\n\n\n\n\n{add_para}'''
		if not os.path.exists(working_directory + f'/articles/{head_line}'):
			os.mkdir(working_directory + f'/articles/{head_line}')
		with open(working_directory + f"/articles/{head_line}/{head_line}.txt", 'w') as file:
			file.write(add_para)

		with open(working_directory + f"/articles/{head_line}/{head_line}.jpg", 'wb') as file:
			file.write(image)

		i += 1

	wb = openpyxl.load_workbook('article.xlsx')
	sheet = wb.active
	row = sheet.max_row
	column = sheet.max_column
	for i in all_data_of_articles:
		sheet.append(i)
	wb.save('article.xlsx')


if __name__ == "__main__":
	if not os.path.exists(working_directory + '/article.xlsx'):
		all_data_of_articles = [['head_line', 'author', 'date', 'image', 'article']]
		wb = openpyxl.Workbook()
		sheet = wb.active
		for i in range(1, 6):
			cell_ = sheet.cell(row=1, column=i)
			cell_.value = all_data_of_articles[0][i-1]
		wb.save(working_directory + "/article.xlsx")
	else:
		all_data_of_articles = [[]]

	if not os.path.exists(working_directory + "/articles/"):
		print("file is not present")
		os.mkdir(working_directory + "/articles")


	main()

	

